import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header_row(ws, headers):
    for col_idx, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

# --- InvoiceLines Sheet ---
ws_inv = wb.active
ws_inv.title = "InvoiceLines"
ws_inv.sheet_properties.tabColor = "2F5496"

invoice_headers = [
    ("Item No", 14),
    ("Description", 35),
    ("Invoice No", 16),
    ("Customer No", 14),
    ("Customer Name", 30),
    ("Quantity", 12),
    ("Unit Price", 14),
    ("Amount", 14),
    ("Sales Order No", 16),
    ("PO Number", 16),
    ("Posting Date", 14),
    ("Requested Delivery", 18),
]
style_header_row(ws_inv, invoice_headers)

# --- OpenSalesOrders Sheet ---
ws_so = wb.create_sheet("OpenSalesOrders")
ws_so.sheet_properties.tabColor = "548235"

so_headers = [
    ("Order No", 14),
    ("Customer No", 14),
    ("Customer Name", 30),
    ("PO Number", 16),
    ("Status", 12),
    ("Requested Delivery", 18),
]
style_header_row(ws_so, so_headers)

output_path = os.path.join(os.path.dirname(__file__), "Route21_Pricing_Report.xlsx")
wb.save(output_path)
print(f"Created: {output_path}")
