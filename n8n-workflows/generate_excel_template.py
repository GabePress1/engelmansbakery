import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# --- PricingData Sheet ---
ws_pricing = wb.active
ws_pricing.title = "PricingData"

pricing_headers = [
    ("Item No.", 14),
    ("Description", 35),
    ("Category", 14),
    ("Unit Cost", 14),
    ("List Price", 14),
    ("Sales Price", 14),
    ("Sales Type", 18),
    ("Sales Code", 16),
    ("Min. Quantity", 14),
    ("UOM", 10),
    ("Price Start Date", 16),
    ("Price End Date", 16),
    ("Inventory", 12),
    ("Blocked", 10),
    ("Last Modified", 20),
    ("Report Date", 14),
]

for col_idx, (header, width) in enumerate(pricing_headers, 1):
    cell = ws_pricing.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws_pricing.column_dimensions[get_column_letter(col_idx)].width = width

ws_pricing.auto_filter.ref = f"A1:{get_column_letter(len(pricing_headers))}1"
ws_pricing.freeze_panes = "A2"
ws_pricing.sheet_properties.tabColor = "2F5496"

# --- DailyLog Sheet ---
ws_log = wb.create_sheet("DailyLog")

log_headers = [
    ("Report Date", 16),
    ("Total Rows", 12),
    ("Unique Items", 14),
    ("Avg List Price", 16),
    ("Blocked Items", 14),
    ("Status", 12),
]

for col_idx, (header, width) in enumerate(log_headers, 1):
    cell = ws_log.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
    ws_log.column_dimensions[get_column_letter(col_idx)].width = width

ws_log.auto_filter.ref = f"A1:{get_column_letter(len(log_headers))}1"
ws_log.freeze_panes = "A2"
ws_log.sheet_properties.tabColor = "548235"

output_path = os.path.join(os.path.dirname(__file__), "Route21_Pricing_Report.xlsx")
wb.save(output_path)
print(f"Created: {output_path}")
